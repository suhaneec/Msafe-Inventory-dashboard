import streamlit as st
import pandas as pd
import numpy as np
import io
from datetime import date
from openpyxl import load_workbook

st.set_page_config(page_title="MSafe Dead Inventory", page_icon="🏗️", layout="wide")

st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=DM+Mono:wght@400;500&family=DM+Sans:wght@300;400;500;600&display=swap');
html, body, [class*="css"] { font-family: 'DM Sans', sans-serif; }
header[data-testid="stHeader"] { display: none; }
.block-container { padding-top: 1.2rem; }
.banner { background: #1E2D3A; color: #fff; padding: 1rem 1.4rem; border-radius: 8px; margin-bottom: 1.2rem; }
.banner h1 { font-size: 1.2rem; font-weight: 600; margin: 0; }
.banner .sub { font-size: 0.75rem; opacity: 0.55; font-family: 'DM Mono', monospace; margin-top: 3px; }
.kpi-row { display: flex; gap: 0.8rem; margin-bottom: 1.2rem; flex-wrap: wrap; }
.kpi { background: #fff; border: 1px solid #E0DBD5; border-radius: 7px; padding: 0.85rem 1.1rem; flex: 1; min-width: 120px; }
.kpi .lbl { font-size: 0.65rem; font-family: 'DM Mono', monospace; text-transform: uppercase; letter-spacing: 0.08em; color: #6A7A88; margin-bottom: 3px; }
.kpi .val { font-size: 1.5rem; font-weight: 600; color: #1E2D3A; line-height: 1; }
.kpi .sub { font-size: 0.68rem; color: #999; margin-top: 3px; }
.kpi.red .val { color: #C84B2F; }
.kpi.amber .val { color: #D97706; }
.sec { font-size: 0.68rem; font-family: 'DM Mono', monospace; text-transform: uppercase;
       letter-spacing: 0.1em; color: #4A5C6A; border-bottom: 1.5px solid #C84B2F;
       padding-bottom: 4px; margin: 1.1rem 0 0.7rem; }
.warn-box { background: #FEF3F0; border-left: 4px solid #C84B2F; border-radius: 0 6px 6px 0;
            padding: 0.6rem 0.9rem; font-size: 0.78rem; color: #7A2E1A; margin-bottom: 0.8rem; }
section[data-testid="stSidebar"] { background: #1E2D3A; }
section[data-testid="stSidebar"] * { color: #B0BEC5 !important; }
section[data-testid="stSidebar"] h2 { color: #fff !important; font-size: 0.9rem; }
</style>
""", unsafe_allow_html=True)


def to_num(v):
    """Safely convert a cell value to float, return 0 if None/blank."""
    if v is None:
        return 0.0
    try:
        return float(v)
    except (ValueError, TypeError):
        return 0.0


def parse_vsoft_file(file_bytes, filename):
    """
    Read Vsoft dead inventory xlsx using openpyxl directly.
    Structure:
      Row 0  : headers  (Item Code | Item Name | Total Qty | Total Amt | 0-60 | | 60-120 | | 120-180 |)
      Row 1  : bucket sub-headers (Qty | Amount | Qty | Amount | Qty | Amount)
      Rows 2+ : data items
      Second-to-last row : totals
      Last row : LOCATION label
    """
    try:
        wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
        ws = wb.active
        all_rows = list(ws.iter_rows(values_only=True))
        wb.close()
    except Exception as e:
        return None, None, f"Cannot open file: {e}"

    if len(all_rows) < 4:
        return None, None, "File has too few rows — check it is the correct Vsoft export."

    # ── Extract location from last rows ───────────────────────────────────────
    location = filename.replace(".xlsx", "").replace(".xls", "")
    for row in reversed(all_rows[-4:]):
        for cell in row:
            if cell and isinstance(cell, str):
                up = cell.upper()
                if any(k in up for k in ["LOCATION", "FACTORY", "GODOWN", "YARD", "DEPOT"]):
                    clean = cell
                    for strip in ["LOCATION _", "LOCATION_", "LOCATION"]:
                        clean = clean.replace(strip, "").replace(strip.lower(), "")
                    location = clean.strip(" _:-")
                    break

    # ── Parse data rows (skip row 0 header, row 1 sub-header, last 2 rows) ───
    data_rows = all_rows[2:-2]

    records = []
    for row in data_rows:
        # Pad row to at least 10 columns
        row = list(row) + [None] * max(0, 10 - len(row))

        item_code = row[0]
        item_name = row[1]

        # Skip blank / totals rows
        if not item_code or str(item_code).strip() == "" or str(item_code).strip().lower() == "none":
            continue

        records.append({
            "Item Code"   : str(item_code).strip(),
            "Item Name"   : str(item_name).strip() if item_name else "",
            "Total Qty"   : to_num(row[2]),
            "Total Amt"   : to_num(row[3]),
            "Qty_0_60"    : to_num(row[4]),
            "Amt_0_60"    : to_num(row[5]),
            "Qty_60_120"  : to_num(row[6]),
            "Amt_60_120"  : to_num(row[7]),
            "Qty_120_180" : to_num(row[8]),
            "Amt_120_180" : to_num(row[9]),
            "Location"    : location,
        })

    if not records:
        return None, None, "No data rows found after parsing."

    df = pd.DataFrame(records)
    # Dead = zero qty in both 0-60 AND 60-120 buckets
    df["Is Dead"] = (df["Qty_0_60"] == 0) & (df["Qty_60_120"] == 0)

    return df, location, None


# ── Sidebar ───────────────────────────────────────────────────────────────────
with st.sidebar:
    st.markdown("## 🏗️ MSafe\nDead Inventory")
    st.markdown("---")
    uploaded = st.file_uploader(
        "Upload location files",
        type=["xlsx"],
        accept_multiple_files=True,
        help="Upload one Vsoft export per location. All files processed together."
    )
    st.markdown("---")
    st.caption(f"MSafe Equipments Pvt Ltd\n{date.today().strftime('%d %b %Y')}")

# ── Banner ────────────────────────────────────────────────────────────────────
st.markdown(f"""
<div class="banner">
  <h1>🏗️ Dead Inventory Intelligence</h1>
  <div class="sub">Items with zero outward movement in 0–120 days · All locations · {date.today().strftime('%d %b %Y')}</div>
</div>
""", unsafe_allow_html=True)

if not uploaded:
    st.markdown("""
    <div style="background:#F7F5F2;border:1.5px dashed #C8BFB5;border-radius:8px;
                padding:2.5rem;text-align:center;color:#888;font-size:0.83rem;margin-top:1rem;">
        <b>Upload your Vsoft location files in the sidebar to begin.</b><br><br>
        In Vsoft → change location to each godown → export dead inventory report → upload all files here at once.<br><br>
        <span style="font-family:monospace;font-size:0.75rem;">One .xlsx file per location</span>
    </div>
    """, unsafe_allow_html=True)
    st.stop()

# ── Parse all uploads ─────────────────────────────────────────────────────────
all_dfs, errors = [], []
for f in uploaded:
    df, loc, err = parse_vsoft_file(f.read(), f.name)
    if err:
        errors.append(f"❌ **{f.name}**: {err}")
    else:
        all_dfs.append(df)

for e in errors:
    st.error(e)

if not all_dfs:
    st.stop()

master = pd.concat(all_dfs, ignore_index=True)
dead   = master[master["Is Dead"]].copy()

# ── Location filter ───────────────────────────────────────────────────────────
all_locs = sorted(master["Location"].unique())
sel_locs = st.sidebar.multiselect("Filter locations", all_locs, default=all_locs)
master_v = master[master["Location"].isin(sel_locs)]
dead_v   = dead[dead["Location"].isin(sel_locs)]

# ── KPIs ──────────────────────────────────────────────────────────────────────
total_items  = len(master_v)
dead_items   = len(dead_v)
dead_pct     = dead_items / total_items * 100 if total_items else 0
dead_qty     = int(dead_v["Total Qty"].sum())
dead_amt     = dead_v["Total Amt"].sum()
total_amt    = master_v["Total Amt"].sum()
dead_amt_pct = dead_amt / total_amt * 100 if total_amt else 0
locs_ct      = master_v["Location"].nunique()

st.markdown(f"""
<div class="kpi-row">
  <div class="kpi red">
    <div class="lbl">Dead SKUs</div>
    <div class="val">{dead_items}</div>
    <div class="sub">{dead_pct:.1f}% of {total_items} lines</div>
  </div>
  <div class="kpi red">
    <div class="lbl">Dead Units</div>
    <div class="val">{dead_qty:,}</div>
    <div class="sub">zero movement 0–120d</div>
  </div>
  <div class="kpi amber">
    <div class="lbl">Dead Value</div>
    <div class="val">₹{dead_amt/100000:.2f}L</div>
    <div class="sub">{dead_amt_pct:.1f}% of total value</div>
  </div>
  <div class="kpi">
    <div class="lbl">Total Inv Value</div>
    <div class="val">₹{total_amt/100000:.2f}L</div>
    <div class="sub">{locs_ct} location(s) loaded</div>
  </div>
</div>
""", unsafe_allow_html=True)

# ── TABS ──────────────────────────────────────────────────────────────────────
tab1, tab2, tab3 = st.tabs(["💀 Dead by Location", "📦 Cross-Location Item View", "📋 Full Inventory"])


# ══════════════════════════════════════════════════════════════
# TAB 1 — Dead by Location
# ══════════════════════════════════════════════════════════════
with tab1:
    if len(dead_v) == 0:
        st.success("No dead stock found across selected locations.")
    else:
        st.markdown('<div class="sec">Location Summary</div>', unsafe_allow_html=True)

        loc_dead = dead_v.groupby("Location").agg(
            Dead_SKUs=("Item Code", "count"),
            Dead_Qty=("Total Qty", "sum"),
            Dead_Value=("Total Amt", "sum"),
        ).reset_index()
        loc_total = master_v.groupby("Location").agg(
            Total_SKUs=("Item Code", "count"),
            Total_Value=("Total Amt", "sum"),
        ).reset_index()
        loc_sum = loc_dead.merge(loc_total, on="Location")
        loc_sum["Dead SKU %"]   = (loc_sum["Dead_SKUs"] / loc_sum["Total_SKUs"] * 100).round(1)
        loc_sum["Dead Value %"] = (loc_sum["Dead_Value"] / loc_sum["Total_Value"] * 100).round(1)
        loc_sum["Dead Value"]   = loc_sum["Dead_Value"].apply(lambda x: f"₹{x:,.0f}")
        loc_sum["RAG"]          = loc_sum["Dead SKU %"].apply(
            lambda x: "🔴" if x >= 40 else ("🟡" if x >= 20 else "🟢")
        )
        loc_sum = loc_sum.sort_values("Dead_Value", ascending=False)

        st.dataframe(
            loc_sum[["RAG","Location","Dead_SKUs","Dead SKU %","Dead_Qty","Dead Value","Dead Value %"]].rename(columns={
                "Dead_SKUs": "Dead SKUs", "Dead_Qty": "Dead Qty (units)"
            }),
            use_container_width=True, hide_index=True
        )

        st.markdown('<div class="sec">Dead Items per Location</div>', unsafe_allow_html=True)

        for loc in sorted(dead_v["Location"].unique()):
            loc_df = dead_v[dead_v["Location"] == loc].sort_values("Total Amt", ascending=False)
            total_val = loc_df["Total Amt"].sum()
            total_qty = int(loc_df["Total Qty"].sum())

            with st.expander(
                f"📍 {loc}  —  {len(loc_df)} dead SKUs  ·  {total_qty} units  ·  ₹{total_val:,.0f}",
                expanded=True
            ):
                show = loc_df[["Item Code","Item Name","Total Qty","Total Amt",
                               "Qty_0_60","Qty_60_120","Qty_120_180"]].copy()
                show["Total Amt"] = show["Total Amt"].apply(lambda x: f"₹{x:,.0f}")
                for c in ["Qty_0_60","Qty_60_120","Qty_120_180"]:
                    show[c] = show[c].apply(lambda x: int(x) if x > 0 else "—")
                show = show.rename(columns={
                    "Total Qty": "Qty", "Total Amt": "Value",
                    "Qty_0_60": "0–60d", "Qty_60_120": "60–120d", "Qty_120_180": "120–180d"
                })
                st.dataframe(show.reset_index(drop=True), use_container_width=True, hide_index=True)


# ══════════════════════════════════════════════════════════════
# TAB 2 — Cross-location analysis
# ══════════════════════════════════════════════════════════════
with tab2:
    st.markdown('<div class="sec">Items Dead Across Multiple Locations</div>', unsafe_allow_html=True)

    if len(dead_v) == 0:
        st.success("No dead items found.")
    else:
        cross = dead_v.groupby(["Item Code","Item Name"]).agg(
            Locations_Dead=("Location", "nunique"),
            Location_Names=("Location", lambda x: ", ".join(sorted(x.unique()))),
            Total_Dead_Qty=("Total Qty", "sum"),
            Total_Dead_Value=("Total Amt", "sum"),
        ).reset_index().sort_values("Total_Dead_Value", ascending=False)

        multi_loc = cross[cross["Locations_Dead"] > 1]
        if len(multi_loc):
            st.markdown(f"""
            <div class="warn-box">
              ⚠️ <b>{len(multi_loc)} SKUs</b> are dead simultaneously across 2+ locations —
              systemic underutilisation. Consider disposal or reallocation.
            </div>
            """, unsafe_allow_html=True)

        cross["RAG"] = cross["Locations_Dead"].apply(
            lambda x: "🔴" if x >= 3 else ("🟡" if x >= 2 else "⚪")
        )
        cross["Total Dead Value"] = cross["Total_Dead_Value"].apply(lambda x: f"₹{x:,.0f}")

        st.dataframe(
            cross[["RAG","Item Code","Item Name","Locations_Dead","Location_Names",
                   "Total_Dead_Qty","Total Dead Value"]].rename(columns={
                "Locations_Dead": "# Locations", "Location_Names": "Found Dead In",
                "Total_Dead_Qty": "Total Dead Qty"
            }),
            use_container_width=True, hide_index=True
        )

        # ── Download ──────────────────────────────────────────────────────────
        st.markdown('<div class="sec">Download Report</div>', unsafe_allow_html=True)
        buf = io.BytesIO()
        with pd.ExcelWriter(buf, engine="openpyxl") as w:
            loc_sum.drop(columns=["RAG"]).to_excel(w, sheet_name="Location Summary", index=False)
            dead_v[[
                "Location","Item Code","Item Name","Total Qty","Total Amt",
                "Qty_0_60","Amt_0_60","Qty_60_120","Amt_60_120","Qty_120_180","Amt_120_180"
            ]].to_excel(w, sheet_name="Dead Items Detail", index=False)
            cross.drop(columns=["RAG","Total Dead Value"]).to_excel(
                w, sheet_name="Cross-Location", index=False
            )
        buf.seek(0)
        st.download_button(
            "⬇️  Download Full Dead Inventory Report (.xlsx)",
            data=buf,
            file_name=f"MSafe_DeadInventory_{date.today()}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )


# ══════════════════════════════════════════════════════════════
# TAB 3 — Full inventory
# ══════════════════════════════════════════════════════════════
with tab3:
    st.markdown('<div class="sec">Complete Inventory — All Items</div>', unsafe_allow_html=True)

    c1, c2 = st.columns([3, 1])
    with c1:
        search = st.text_input("Search item name or code", placeholder="e.g. Prop jack, MSF5SF…")
    with c2:
        show_filter = st.selectbox("Show", ["All", "Dead only", "Active only"])

    view_df = master_v.copy()
    if search:
        mask = (
            view_df["Item Name"].str.contains(search, case=False, na=False) |
            view_df["Item Code"].str.contains(search, case=False, na=False)
        )
        view_df = view_df[mask]
    if show_filter == "Dead only":
        view_df = view_df[view_df["Is Dead"]]
    elif show_filter == "Active only":
        view_df = view_df[~view_df["Is Dead"]]

    view_df = view_df.copy()
    view_df["Status"] = view_df["Is Dead"].apply(lambda x: "🔴 Dead" if x else "🟢 Active")
    view_df["Value"]  = view_df["Total Amt"].apply(lambda x: f"₹{x:,.0f}")
    for c in ["Qty_0_60","Qty_60_120","Qty_120_180"]:
        view_df[c] = view_df[c].apply(lambda x: int(x) if x > 0 else "—")

    st.dataframe(
        view_df[[
            "Status","Location","Item Code","Item Name","Total Qty","Value",
            "Qty_0_60","Qty_60_120","Qty_120_180"
        ]].rename(columns={
            "Total Qty": "Qty", "Qty_0_60": "0–60d",
            "Qty_60_120": "60–120d", "Qty_120_180": "120–180d"
        }).sort_values(["Location","Status"]).reset_index(drop=True),
        use_container_width=True, hide_index=True
    )
